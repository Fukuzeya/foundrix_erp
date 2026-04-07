"""Data importer for parsing CSV and Excel files into structured records.

Modules provide a field mapping and validation schema. The importer
handles parsing, validation, and error reporting per row.

Usage::

    importer = DataImporter()
    result = await importer.from_csv(
        file_content=csv_bytes,
        columns=["name", "email", "phone"],
        required=["name", "email"],
        validators={"email": validate_email},
    )

    if result.errors:
        # Return errors to the user
        ...
    else:
        # Process result.rows
        ...
"""

import csv
import io
import logging
from dataclasses import dataclass, field
from typing import Any, Callable

logger = logging.getLogger(__name__)


@dataclass
class ImportError:
    """A single validation error in an import row."""

    row: int
    column: str
    message: str


@dataclass
class ImportResult:
    """Result of a data import operation."""

    rows: list[dict[str, Any]] = field(default_factory=list)
    errors: list[ImportError] = field(default_factory=list)
    total_rows: int = 0
    success_count: int = 0
    error_count: int = 0

    @property
    def has_errors(self) -> bool:
        return len(self.errors) > 0


class DataImporter:
    """Import data from CSV or Excel files."""

    async def from_csv(
        self,
        file_content: bytes,
        columns: list[str],
        *,
        required: list[str] | None = None,
        validators: dict[str, Callable[[str], bool]] | None = None,
        max_rows: int = 10000,
    ) -> ImportResult:
        """Parse and validate a CSV file.

        Args:
            file_content: Raw CSV bytes.
            columns: Expected column names (used for mapping).
            required: Columns that must have non-empty values.
            validators: Dict of column → validation function.
            max_rows: Maximum rows to process (safety limit).

        Returns:
            ImportResult with parsed rows and any validation errors.
        """
        required = required or []
        validators = validators or {}
        result = ImportResult()

        text_content = file_content.decode("utf-8-sig")  # Handle BOM
        reader = csv.DictReader(io.StringIO(text_content))

        for row_num, row in enumerate(reader, start=2):  # Row 1 is header
            if row_num - 1 > max_rows:
                result.errors.append(
                    ImportError(row=row_num, column="", message=f"Exceeded max rows ({max_rows})")
                )
                break

            result.total_rows += 1
            row_data: dict[str, Any] = {}
            row_valid = True

            for col in columns:
                value = (row.get(col) or "").strip()
                row_data[col] = value

                # Required check
                if col in required and not value:
                    result.errors.append(
                        ImportError(row=row_num, column=col, message=f"Required field '{col}' is empty")
                    )
                    row_valid = False

                # Custom validation
                if value and col in validators:
                    try:
                        if not validators[col](value):
                            result.errors.append(
                                ImportError(row=row_num, column=col, message=f"Invalid value for '{col}'")
                            )
                            row_valid = False
                    except Exception as e:
                        result.errors.append(
                            ImportError(row=row_num, column=col, message=str(e))
                        )
                        row_valid = False

            if row_valid:
                result.rows.append(row_data)
                result.success_count += 1
            else:
                result.error_count += 1

        return result

    async def from_excel(
        self,
        file_content: bytes,
        columns: list[str],
        *,
        required: list[str] | None = None,
        validators: dict[str, Callable[[str], bool]] | None = None,
        sheet_name: str | None = None,
        max_rows: int = 10000,
    ) -> ImportResult:
        """Parse and validate an Excel file.

        Requires openpyxl. Converts to CSV internally for consistent parsing.
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(file_content), read_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active

            output = io.StringIO()
            writer = csv.writer(output)

            for row in ws.iter_rows(values_only=True):
                writer.writerow([str(cell) if cell is not None else "" for cell in row])

            csv_bytes = output.getvalue().encode("utf-8")
            return await self.from_csv(
                csv_bytes, columns, required=required, validators=validators, max_rows=max_rows
            )

        except ImportError:
            result = ImportResult()
            result.errors.append(
                ImportError(row=0, column="", message="openpyxl not installed for Excel support")
            )
            return result
